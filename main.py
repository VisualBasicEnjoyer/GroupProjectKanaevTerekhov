import sys
import openpyxl
import datetime
from PyQt6 import QtWidgets
from mainwindow import Ui_MainWindow
from windowlistadd import Ui_DialogListAdd
from windowlistadd2 import Ui_DialogListAdd2
from windowlistadd3 import Ui_DialogListAdd3
from windowlistadd4 import Ui_DialogListAdd4

def main():

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

    wb = openpyxl.load_workbook('table.xlsx')
    wb.iso_dates = True

    sheets = wb.sheetnames

    def listupdate():
        sheet = wb[sheets[0]]
        rowcount = 1
        while True:
            if sheet.cell(rowcount, 1).value:
                rowcount += 1
            else:
                break

        rowc = 0
        colc = 0
        for i in range((rowcount - 1) * 6):
            ui.table_list.item(rowc, colc).setText(str(sheet.cell(rowc + 1, colc + 1).value))
            colc += 1
            if colc == 6:
                colc = 0
                rowc += 1

    def accountupdate():
        sheet = wb[sheets[1]]
        rowcount = 1
        while True:
            if sheet.cell(rowcount, 1).value:
                rowcount += 1
            else:
                break

        rowc = 0
        colc = 0
        for i in range((rowcount - 1) * 5):
            ui.table_account.item(rowc, colc).setText(str(sheet.cell(rowc + 1, colc + 1).value))
            colc += 1
            if colc == 5:
                colc = 0
                rowc += 1

        total = 0
        for i in range(rowcount - 1):
            total += int(wb[sheets[1]].cell(i + 1, 5).value)
        ui.total_text.setText(str(total))

        quantity = 0
        for i in range(rowcount - 1):
            quantity += int(wb[sheets[1]].cell(i + 1, 3).value)
        ui.quantity_text.setText(str(quantity))

    accountupdate()
    listupdate()

    def openDialogList():
        global DialogListAdd
        DialogListAdd = QtWidgets.QDialog()
        ui = Ui_DialogListAdd()
        ui.setupUi(DialogListAdd)
        DialogListAdd.show()
        ui.warn_label.hide()
        ui.warn_label2.hide()

        def btn_list_add1_click():
            if not ui.number_text.text() or not ui.name_text.text() or not ui.quantity_text.text() or not ui.manufacturer_text.text() or not ui.purchase_price_text.text() or not ui.sale_price_text.text():
                ui.warn_label2.hide()
                ui.warn_label.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text.text()):
                        rowcount += 1
                    else:
                        ui.warn_label.hide()
                        ui.warn_label2.show()
                        return

                sheet = wb[sheets[0]]
                rowcount = 1
                while True:
                    if sheet.cell(rowcount, 1).value:
                        rowcount += 1
                    else:
                        break
                sheet.cell(rowcount, 1).value = int(ui.number_text.text())
                sheet.cell(rowcount, 2).value = str(ui.name_text.text())
                sheet.cell(rowcount, 3).value = int(ui.quantity_text.text())
                sheet.cell(rowcount, 4).value = str(ui.manufacturer_text.text())
                sheet.cell(rowcount, 5).value = int(ui.purchase_price_text.text())
                sheet.cell(rowcount, 6).value = int(ui.sale_price_text.text())
                wb.save('table.xlsx')
                listupdate()
                DialogListAdd.close()


        ui.btn_list_add1.clicked.connect(btn_list_add1_click)

    ui.btn_list.clicked.connect(openDialogList)

    def openDialogList3():
        global DialogListAdd3
        DialogListAdd3 = QtWidgets.QDialog()
        ui = Ui_DialogListAdd3()
        ui.setupUi(DialogListAdd3)
        DialogListAdd3.show()
        ui.warn_label1_3.hide()
        ui.warn_label2_3.hide()

        def btn_list_delete3_click():

            sheet = wb[sheets[0]]
            rowcountglob = 1
            while True:
                if sheet.cell(rowcountglob, 1).value:
                    rowcountglob += 1
                else:
                    break

            if not ui.number_text_3.text():
                ui.warn_label2_3.hide()
                ui.warn_label1_3.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text_3.text()):
                        rowcount += 1
                    else:
                        for i in range(6):
                            sheet.cell(rowcount, i + 1).value = ""

                        for i in range(rowcountglob-rowcount):
                            for u in range(6):
                                sheet.cell(rowcount + i, u + 1).value = sheet.cell(rowcount + 1 + i, u + 1).value

                        for i in range(6):
                            sheet.cell(rowcountglob-1, i + 1).value = ""

                        wb.save('table.xlsx')
                        listupdate()
                        DialogListAdd3.close()

                ui.warn_label1_3.hide()
                ui.warn_label2_3.show()
                return

        ui.btn_list_delete3.clicked.connect(btn_list_delete3_click)

    ui.btn_list_4.clicked.connect(openDialogList3)

    def openDialogList4():
        global DialogListAdd4
        DialogListAdd4 = QtWidgets.QDialog()
        ui = Ui_DialogListAdd4()
        ui.setupUi(DialogListAdd4)
        DialogListAdd4.show()
        ui.warn_label1_4.hide()
        ui.warn_label2_4.hide()

        ui.name_label_4.hide()
        ui.name_text_4.hide()
        ui.quantity_label_4.hide()
        ui.quantity_text_4.hide()
        ui.manufacturer_label_4.hide()
        ui.manufacturer_text_4.hide()
        ui.purchase_price_label_4.hide()
        ui.purchase_price_text_4.hide()
        ui.sale_price_label_4.hide()
        ui.sale_price_text_4.hide()

        ui.btn_list_edit2_4.hide()

        def btn_list_edit_4_click():
            if not ui.number_text_4.text():
                ui.warn_label2_4.hide()
                ui.warn_label1_4.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text_4.text()):
                        rowcount += 1
                    else:
                        ui.warn_label1_4.hide()
                        ui.warn_label2_4.hide()

                        ui.name_label_4.show()
                        ui.name_text_4.show()
                        ui.quantity_label_4.show()
                        ui.quantity_text_4.show()
                        ui.manufacturer_label_4.show()
                        ui.manufacturer_text_4.show()
                        ui.purchase_price_label_4.show()
                        ui.purchase_price_text_4.show()
                        ui.sale_price_label_4.show()
                        ui.sale_price_text_4.show()

                        ui.name_text_4.setText(str(sheet.cell(rowcount, 2).value))
                        ui.quantity_text_4.setText(str(sheet.cell(rowcount, 3).value))
                        ui.manufacturer_text_4.setText(str(sheet.cell(rowcount, 4).value))
                        ui.purchase_price_text_4.setText(str(sheet.cell(rowcount, 5).value))
                        ui.sale_price_text_4.setText(str(sheet.cell(rowcount, 6).value))

                        ui.number_text_4.setReadOnly(True)

                        ui.btn_list_edit_4.hide()
                        ui.btn_list_edit2_4.show()
                        return

                ui.warn_label1_4.hide()
                ui.warn_label2_4.show()
                return

        ui.btn_list_edit_4.clicked.connect(btn_list_edit_4_click)

        def btn_list_edit2_4_click():

            if not ui.name_text_4.text() or not ui.quantity_text_4.text() or not ui.manufacturer_text_4.text() or not ui.purchase_price_text_4.text() or not ui.sale_price_text_4.text():
                ui.warn_label2_4.hide()
                ui.warn_label1_4.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text_4.text()):
                        rowcount += 1
                    else:
                        sheet.cell(rowcount, 2).value = str(ui.name_text_4.text())
                        sheet.cell(rowcount, 3).value = int(ui.quantity_text_4.text())
                        sheet.cell(rowcount, 4).value = str(ui.manufacturer_text_4.text())
                        sheet.cell(rowcount, 5).value = int(ui.purchase_price_text_4.text())
                        sheet.cell(rowcount, 6).value = int(ui.sale_price_text_4.text())
                        wb.save('table.xlsx')
                        listupdate()
                        DialogListAdd4.close()

        ui.btn_list_edit2_4.clicked.connect(btn_list_edit2_4_click)

    ui.btn_list_3.clicked.connect(openDialogList4)

    def openDialogList2():
        global DialogListAdd2
        DialogListAdd2 = QtWidgets.QDialog()
        ui = Ui_DialogListAdd2()
        ui.setupUi(DialogListAdd2)
        DialogListAdd2.show()
        ui.warn_label1_2.hide()
        ui.warn_label2_2.hide()
        ui.warn_label3_2.hide()
        ui.btn_list_cart2_2.hide()
        ui.quantity_text_2.hide()
        ui.quantity_label_2.hide()

        def btn_list_cart2_click():
            if not ui.number_text_2.text():
                ui.warn_label2_2.hide()
                ui.warn_label1_2.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text_2.text()):
                        rowcount += 1
                    else:
                        ui.warn_label1_2.hide()
                        ui.warn_label2_2.hide()

                        ui.number_text_2.setReadOnly(True)
                        ui.quantity_text_2.show()
                        ui.quantity_label_2.show()

                        ui.btn_list_cart2.hide()
                        ui.btn_list_cart2_2.show()
                        return

                ui.warn_label1_2.hide()
                ui.warn_label2_2.show()
                return

        ui.btn_list_cart2.clicked.connect(btn_list_cart2_click)

        def btn_list_cart2_2_click():
            if not ui.quantity_text_2.text():
                ui.warn_label1_2.show()
            else:
                sheet = wb[sheets[0]]
                rowcount = 1
                for i in range(200):
                    if str(sheet.cell(rowcount, 1).value) != str(ui.number_text_2.text()):
                        rowcount += 1
                    else:
                        break
                if (wb[sheets[0]].cell(rowcount, 3).value - int(ui.quantity_text_2.text())) < 0:
                    ui.warn_label1_2.hide()
                    ui.warn_label3_2.show()
                    return
                else:
                    rowcount2 = 1
                    while True:
                        if wb[sheets[1]].cell(rowcount2, 1).value:
                            rowcount2 += 1
                        else:
                            break
                    wb[sheets[1]].cell(rowcount2, 1).value = wb[sheets[0]].cell(rowcount, 1).value
                    wb[sheets[1]].cell(rowcount2, 2).value = wb[sheets[0]].cell(rowcount, 2).value
                    wb[sheets[1]].cell(rowcount2, 3).value = int(ui.quantity_text_2.text())
                    wb[sheets[0]].cell(rowcount, 3).value -= int(ui.quantity_text_2.text())
                    wb[sheets[1]].cell(rowcount2, 4).value = wb[sheets[0]].cell(rowcount, 4).value

                    wb[sheets[1]].cell(rowcount2, 6).value = wb[sheets[0]].cell(rowcount, 5).value
                    wb[sheets[1]].cell(rowcount2, 7).value = datetime.date.today()
                    wb[sheets[1]].cell(rowcount2, 8).value = wb[sheets[0]].cell(rowcount, 6).value

                    wb[sheets[1]].cell(rowcount2, 5).value = int(ui.quantity_text_2.text()) * int(wb[sheets[0]].cell(rowcount, 6).value)
                    wb.save('table.xlsx')
                    listupdate()
                    accountupdate()
                    DialogListAdd2.close()

        ui.btn_list_cart2_2.clicked.connect(btn_list_cart2_2_click)

    ui.btn_list_2.clicked.connect(openDialogList2)

    def btn_account_2_click():
        rowcount2 = 1
        while True:
            if wb[sheets[1]].cell(rowcount2, 1).value:
                rowcount2 += 1
            else:
                break

        for u in range(rowcount2-1):
            rowcount = 1
            for i in range(200):
                if str(wb[sheets[0]].cell(rowcount, 1).value) != str(wb[sheets[1]].cell(u+1, 1).value):
                    rowcount += 1
                else:
                    wb[sheets[0]].cell(rowcount, 3).value = wb[sheets[1]].cell(u+1, 3).value + wb[sheets[0]].cell(rowcount, 3).value
                    for y in range(8):
                        wb[sheets[1]].cell(u+1, y+1).value = ""
            for y in range(8):
                wb[sheets[1]].cell(u + 1, y + 1).value = ""
        wb.save('table.xlsx')
        listupdate()
        accountupdate()

    ui.btn_account_2.clicked.connect(btn_account_2_click)

    def btn_account1_click():
        rowcount2 = 1
        while True:
            if wb[sheets[1]].cell(rowcount2, 1).value:
                rowcount2 += 1
            else:
                break

        rowcount3 = 1
        rowc = 1
        while True:
            if wb[sheets[2]].cell(rowcount3, 1).value:
                rowcount3 += 1
            else:
                break

        for i in range(rowcount2 - 1):
            wb[sheets[2]].cell(rowcount3, 1).value = wb[sheets[1]].cell(rowc, 1).value
            wb[sheets[2]].cell(rowcount3, 2).value = wb[sheets[1]].cell(rowc, 2).value
            wb[sheets[2]].cell(rowcount3, 3).value = wb[sheets[1]].cell(rowc, 3).value
            wb[sheets[2]].cell(rowcount3, 4).value = wb[sheets[1]].cell(rowc, 4).value
            wb[sheets[2]].cell(rowcount3, 5).value = wb[sheets[1]].cell(rowc, 5).value
            wb[sheets[2]].cell(rowcount3, 6).value = wb[sheets[1]].cell(rowc, 6).value
            wb[sheets[2]].cell(rowcount3, 7).value = datetime.date.today()
            wb[sheets[2]].cell(rowcount3, 8).value = wb[sheets[1]].cell(rowc, 8).value
            for y in range(8):
                wb[sheets[1]].cell(i + 1, y + 1).value = ""
            rowc += 1
            rowcount3 += 1

        wb.save('table.xlsx')
        listupdate()
        accountupdate()

    ui.btn_account1.clicked.connect(btn_account1_click)

    sys.exit(app.exec())

if __name__ == "__main__":
    main()