import openpyxl
import datetime

wb = openpyxl.load_workbook('table.xlsx')
sheets = wb.sheetnames
rowcount2 = 1
print(wb[sheets[2]].cell(1, 7).value)
print(datetime.date.today())